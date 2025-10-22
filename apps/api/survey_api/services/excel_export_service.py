"""
Excel and CSV export service for survey data.

This service handles the generation of Excel and CSV files for calculated
and interpolated survey data, including proper formatting, metadata sheets,
and file naming conventions.
"""
import io
import re
from datetime import datetime
from typing import BinaryIO, Literal, Tuple

import pandas as pd
from django.core.exceptions import ObjectDoesNotExist
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from survey_api.models import CalculatedSurvey, InterpolatedSurvey, ComparisonResult


class ExcelExportService:
    """Service for exporting survey data to Excel and CSV formats."""

    # Column definitions for calculated survey data
    CALCULATED_COLUMNS = [
        'MD (m)', 'Inc (deg)', 'Azi (deg)',
        'Easting (m)', 'Northing (m)', 'TVD (m)',
        'DLS (deg/30m)', 'Build Rate (deg/30m)', 'Turn Rate (deg/30m)'
    ]

    # Column definitions for interpolated survey data
    INTERPOLATED_COLUMNS = [
        'MD (m)', 'Inc (deg)', 'Azi (deg)',
        'Easting (m)', 'Northing (m)', 'TVD (m)'
    ]

    # Excel styling constants
    HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
    HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center')

    METADATA_LABEL_FONT = Font(bold=True, size=10)
    METADATA_VALUE_ALIGNMENT = Alignment(horizontal='left', vertical='center')

    @classmethod
    def export_calculated_survey(
        cls,
        calculated_survey_id: str,
        format: Literal['excel', 'csv'] = 'excel'
    ) -> Tuple[BinaryIO, str, str]:
        """
        Export calculated survey data to Excel or CSV format.

        Args:
            calculated_survey_id: UUID of the CalculatedSurvey instance
            format: Export format ('excel' or 'csv')

        Returns:
            Tuple of (file_buffer, filename, content_type)

        Raises:
            ObjectDoesNotExist: If calculated survey not found
            ValueError: If data is invalid or incomplete
        """
        try:
            # Fetch calculated survey with related data
            calculated = CalculatedSurvey.objects.select_related(
                'survey_data',
                'survey_data__survey_file',
                'survey_data__survey_file__run',
                'survey_data__survey_file__run__location',
                'survey_data__survey_file__run__depth',
                'survey_data__survey_file__run__tieon'
            ).get(id=calculated_survey_id)

            # Validate data is available
            if calculated.calculation_status not in ['calculated', 'completed']:
                raise ValueError(
                    f"Cannot export survey with status '{calculated.calculation_status}'. "
                    "Only calculated/completed calculations can be exported."
                )

            if not calculated.survey_data:
                raise ValueError("Survey data is missing from calculated survey.")

            # Generate file based on format
            if format == 'excel':
                return cls._export_calculated_excel(calculated)
            else:
                return cls._export_calculated_csv(calculated)

        except CalculatedSurvey.DoesNotExist:
            raise ObjectDoesNotExist(
                f"CalculatedSurvey with id '{calculated_survey_id}' does not exist."
            )

    @classmethod
    def export_interpolated_survey(
        cls,
        interpolated_survey_id: str,
        format: Literal['excel', 'csv'] = 'excel'
    ) -> Tuple[BinaryIO, str, str]:
        """
        Export interpolated survey data to Excel or CSV format.

        Args:
            interpolated_survey_id: UUID of the InterpolatedSurvey instance
            format: Export format ('excel' or 'csv')

        Returns:
            Tuple of (file_buffer, filename, content_type)

        Raises:
            ObjectDoesNotExist: If interpolated survey not found
            ValueError: If data is invalid or incomplete
        """
        try:
            # Fetch interpolated survey with related data
            interpolated = InterpolatedSurvey.objects.select_related(
                'calculated_survey',
                'calculated_survey__survey_data',
                'calculated_survey__survey_data__survey_file',
                'calculated_survey__survey_data__survey_file__run',
                'calculated_survey__survey_data__survey_file__run__location',
                'calculated_survey__survey_data__survey_file__run__depth',
                'calculated_survey__survey_data__survey_file__run__tieon'
            ).get(id=interpolated_survey_id)

            # Validate data is available
            if interpolated.interpolation_status != 'completed':
                raise ValueError(
                    f"Cannot export survey with status '{interpolated.interpolation_status}'. "
                    "Only completed interpolations can be exported."
                )

            # Generate file based on format
            if format == 'excel':
                return cls._export_interpolated_excel(interpolated)
            else:
                return cls._export_interpolated_csv(interpolated)

        except InterpolatedSurvey.DoesNotExist:
            raise ObjectDoesNotExist(
                f"InterpolatedSurvey with id '{interpolated_survey_id}' does not exist."
            )

    @classmethod
    def _export_calculated_excel(cls, calculated: CalculatedSurvey) -> Tuple[BinaryIO, str, str]:
        """Generate Excel file for calculated survey data."""
        wb = Workbook()
        run = calculated.survey_data.survey_file.run

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Create metadata sheet
        cls._create_metadata_sheet(wb, run, 'calculated')

        # Create survey data sheet
        ws_data = wb.create_sheet('Survey Data')
        cls._write_calculated_data(ws_data, calculated)

        # Generate filename
        filename = cls._generate_filename(run.run_name, 'calculated', 'excel')

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer, filename, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    @classmethod
    def _export_interpolated_excel(cls, interpolated: InterpolatedSurvey) -> Tuple[BinaryIO, str, str]:
        """Generate Excel file for interpolated survey data."""
        wb = Workbook()
        run = interpolated.calculated_survey.survey_data.survey_file.run

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Create metadata sheet
        cls._create_metadata_sheet(wb, run, 'interpolated', interpolated)

        # Create survey data sheet
        ws_data = wb.create_sheet('Survey Data')
        cls._write_interpolated_data(ws_data, interpolated)

        # Generate filename
        filename = cls._generate_filename(
            run.run_name,
            'interpolated',
            'excel'
        )

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer, filename, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    @classmethod
    def _create_metadata_sheet(
        cls,
        wb: Workbook,
        run,
        data_type: Literal['calculated', 'interpolated'],
        interpolated: InterpolatedSurvey = None
    ) -> None:
        """Create metadata sheet with run information."""
        ws = wb.create_sheet('Metadata', 0)

        # Title
        ws['A1'] = 'Survey Export Metadata'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')

        # Run information
        row = 3
        metadata = [
            ('Run Name', str(run)),
            ('Export Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ('Data Type', data_type.capitalize()),
        ]

        # Add location info if available
        if hasattr(run, 'location') and run.location:
            metadata.extend([
                ('Latitude', f"{run.location.latitude}"),
                ('Longitude', f"{run.location.longitude}"),
            ])

        # Add depth info if available
        if hasattr(run, 'depth') and run.depth:
            metadata.extend([
                ('Reference Height (m)', f"{run.depth.reference_height:.2f}"),
                ('Reference Elevation (m)', f"{run.depth.reference_elevation:.2f}"),
            ])

        # Add tie-on info if available
        if hasattr(run, 'tieon') and run.tieon:
            metadata.extend([
                ('Tie-On MD (m)', f"{run.tieon.md:.2f}"),
                ('Tie-On Inclination (deg)', f"{run.tieon.inc:.2f}"),
                ('Tie-On Azimuth (deg)', f"{run.tieon.azi:.2f}"),
            ])

        # Add interpolation info if applicable
        if data_type == 'interpolated' and interpolated:
            metadata.extend([
                ('Resolution (m)', str(interpolated.resolution)),
                ('Point Count', str(interpolated.point_count)),
            ])

        # Write metadata rows
        for label, value in metadata:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = cls.METADATA_LABEL_FONT
            ws[f'B{row}'] = value
            ws[f'B{row}'].alignment = cls.METADATA_VALUE_ALIGNMENT
            row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40

    @classmethod
    def _write_calculated_data(cls, ws, calculated: CalculatedSurvey) -> None:
        """Write calculated survey data to worksheet."""
        # Prepare data
        survey_data = calculated.survey_data
        row_count = survey_data.row_count

        # Create DataFrame
        data = {
            'MD (m)': survey_data.md_data[:row_count],
            'Inc (deg)': survey_data.inc_data[:row_count],
            'Azi (deg)': survey_data.azi_data[:row_count],
            'Easting (m)': calculated.easting[:row_count],
            'Northing (m)': calculated.northing[:row_count],
            'TVD (m)': calculated.tvd[:row_count],
            'DLS (deg/30m)': calculated.dls[:row_count],
            'Build Rate (deg/30m)': calculated.build_rate[:row_count],
            'Turn Rate (deg/30m)': calculated.turn_rate[:row_count],
        }
        df = pd.DataFrame(data)

        # Write headers
        for col_num, column_name in enumerate(cls.CALCULATED_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_num, value=column_name)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.alignment = cls.HEADER_ALIGNMENT

        # Write data rows
        for row_num, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                # Format numbers to 2 decimal places
                if isinstance(value, (int, float)):
                    cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal='right')

        # Auto-adjust column widths
        for col_num in range(1, len(cls.CALCULATED_COLUMNS) + 1):
            ws.column_dimensions[chr(64 + col_num)].width = 15

    @classmethod
    def _write_interpolated_data(cls, ws, interpolated: InterpolatedSurvey) -> None:
        """Write interpolated survey data to worksheet."""
        # Prepare data
        row_count = interpolated.point_count

        # Create DataFrame
        data = {
            'MD (m)': interpolated.md_interpolated[:row_count],
            'Inc (deg)': interpolated.inc_interpolated[:row_count],
            'Azi (deg)': interpolated.azi_interpolated[:row_count],
            'Easting (m)': interpolated.easting_interpolated[:row_count],
            'Northing (m)': interpolated.northing_interpolated[:row_count],
            'TVD (m)': interpolated.tvd_interpolated[:row_count],
        }
        df = pd.DataFrame(data)

        # Write headers
        for col_num, column_name in enumerate(cls.INTERPOLATED_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_num, value=column_name)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.alignment = cls.HEADER_ALIGNMENT

        # Write data rows
        for row_num, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                # Format numbers to 2 decimal places
                if isinstance(value, (int, float)):
                    cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal='right')

        # Auto-adjust column widths
        for col_num in range(1, len(cls.INTERPOLATED_COLUMNS) + 1):
            ws.column_dimensions[chr(64 + col_num)].width = 15

    @classmethod
    def _export_calculated_csv(cls, calculated: CalculatedSurvey) -> Tuple[BinaryIO, str, str]:
        """Generate CSV file for calculated survey data."""
        # Prepare data
        survey_data = calculated.survey_data
        row_count = survey_data.row_count
        run = survey_data.survey_file.run

        data = {
            'MD (m)': survey_data.md_data[:row_count],
            'Inc (deg)': survey_data.inc_data[:row_count],
            'Azi (deg)': survey_data.azi_data[:row_count],
            'Easting (m)': calculated.easting[:row_count],
            'Northing (m)': calculated.northing[:row_count],
            'TVD (m)': calculated.tvd[:row_count],
            'DLS (deg/30m)': calculated.dls[:row_count],
            'Build Rate (deg/30m)': calculated.build_rate[:row_count],
            'Turn Rate (deg/30m)': calculated.turn_rate[:row_count],
        }
        df = pd.DataFrame(data)

        # Create buffer with metadata header
        buffer = io.StringIO()

        # Write metadata as comments
        buffer.write(f"# Survey Export - Calculated Data\n")
        buffer.write(f"# Run Name: {str(run)}\n")
        buffer.write(f"# Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        buffer.write(f"# Row Count: {row_count}\n")
        buffer.write("#\n")

        # Write CSV data
        df.to_csv(buffer, index=False, float_format='%.2f')

        # Convert to bytes
        bytes_buffer = io.BytesIO(buffer.getvalue().encode('utf-8'))
        bytes_buffer.seek(0)

        # Generate filename
        filename = cls._generate_filename(run.run_name, 'calculated', 'csv')

        return bytes_buffer, filename, 'text/csv'

    @classmethod
    def _export_interpolated_csv(cls, interpolated: InterpolatedSurvey) -> Tuple[BinaryIO, str, str]:
        """Generate CSV file for interpolated survey data."""
        # Prepare data
        row_count = interpolated.point_count
        run = interpolated.calculated_survey.survey_data.survey_file.run

        data = {
            'MD (m)': interpolated.md_interpolated[:row_count],
            'Inc (deg)': interpolated.inc_interpolated[:row_count],
            'Azi (deg)': interpolated.azi_interpolated[:row_count],
            'Easting (m)': interpolated.easting_interpolated[:row_count],
            'Northing (m)': interpolated.northing_interpolated[:row_count],
            'TVD (m)': interpolated.tvd_interpolated[:row_count],
        }
        df = pd.DataFrame(data)

        # Create buffer with metadata header
        buffer = io.StringIO()

        # Write metadata as comments
        buffer.write(f"# Survey Export - Interpolated Data\n")
        buffer.write(f"# Run Name: {str(run)}\n")
        buffer.write(f"# Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        buffer.write(f"# Resolution: {interpolated.resolution} m\n")
        buffer.write(f"# Point Count: {row_count}\n")
        buffer.write("#\n")

        # Write CSV data
        df.to_csv(buffer, index=False, float_format='%.2f')

        # Convert to bytes
        bytes_buffer = io.BytesIO(buffer.getvalue().encode('utf-8'))
        bytes_buffer.seek(0)

        # Generate filename
        filename = cls._generate_filename(
            run.run_name,
            'interpolated',
            'csv'
        )

        return bytes_buffer, filename, 'text/csv'

    @classmethod
    def _generate_filename(
        cls,
        run_name: str,
        data_type: Literal['calculated', 'interpolated'],
        format: Literal['excel', 'csv']
    ) -> str:
        """
        Generate standardized filename for export.

        Format: {run_name}_{data_type}_{timestamp}.{ext}

        Args:
            run_name: Name of the run
            data_type: Type of data (calculated or interpolated)
            format: File format (excel or csv)

        Returns:
            Sanitized filename with proper extension
        """
        # Sanitize run name (remove special characters, limit length)
        sanitized_name = re.sub(r'[^\w\s-]', '', run_name).strip()
        sanitized_name = re.sub(r'[-\s]+', '_', sanitized_name)

        # Limit length to 50 characters
        if len(sanitized_name) > 50:
            sanitized_name = sanitized_name[:50]

        # Generate timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        # Determine extension
        ext = 'xlsx' if format == 'excel' else 'csv'

        # Construct filename
        filename = f"{sanitized_name}_{data_type}_{timestamp}.{ext}"

        return filename

    @classmethod
    def export_comparison_results(
        cls,
        comparison_id: str,
        format: Literal['excel', 'csv'] = 'excel'
    ) -> Tuple[BinaryIO, str, str]:
        """
        Export comparison results to Excel or CSV format.

        Args:
            comparison_id: UUID of the ComparisonResult instance
            format: Export format ('excel' or 'csv')

        Returns:
            Tuple of (file_buffer, filename, content_type)

        Raises:
            ObjectDoesNotExist: If comparison not found
            ValueError: If data is invalid or incomplete
        """
        try:
            # Fetch comparison with related data
            comparison = ComparisonResult.objects.select_related(
                'run',
                'primary_survey',
                'reference_survey',
                'primary_survey__survey_file',
                'reference_survey__survey_file'
            ).get(id=comparison_id)

            # Generate file based on format
            if format == 'excel':
                return cls._export_comparison_excel(comparison)
            else:
                return cls._export_comparison_csv(comparison)

        except ComparisonResult.DoesNotExist:
            raise ObjectDoesNotExist(
                f"ComparisonResult with id '{comparison_id}' does not exist."
            )

    @classmethod
    def _export_comparison_excel(cls, comparison: ComparisonResult) -> Tuple[BinaryIO, str, str]:
        """Generate Excel file for comparison results with multiple sheets."""
        wb = Workbook()
        run = comparison.run

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Create Summary sheet
        ws_summary = wb.create_sheet('Summary', 0)
        cls._write_comparison_summary(ws_summary, comparison)

        # Create Position Deltas sheet
        ws_position = wb.create_sheet('Position Deltas')
        cls._write_position_deltas(ws_position, comparison)

        # Create Angular Deltas sheet
        ws_angular = wb.create_sheet('Angular Deltas')
        cls._write_angular_deltas(ws_angular, comparison)

        # Create Reference Data sheet
        ws_reference = wb.create_sheet('Reference Data')
        cls._write_survey_reference_data(ws_reference, comparison, 'reference')

        # Create Comparison Data sheet
        ws_comparison = wb.create_sheet('Comparison Data')
        cls._write_survey_reference_data(ws_comparison, comparison, 'primary')

        # Generate filename
        primary_file = comparison.primary_survey.survey_file
        reference_file = comparison.reference_survey.survey_file
        filename = cls._generate_comparison_filename(
            run.run_name,
            primary_file.file_name if primary_file else 'primary',
            reference_file.file_name if reference_file else 'reference',
            format='excel'
        )

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer, filename, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    @classmethod
    def _write_comparison_summary(cls, ws, comparison: ComparisonResult) -> None:
        """Write comparison summary with statistics."""
        # Title
        ws['A1'] = 'Survey Comparison Summary'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')

        row = 3

        # Basic info
        primary_file = comparison.primary_survey.survey_file
        reference_file = comparison.reference_survey.survey_file

        metadata = [
            ('Run Name', str(comparison.run)),
            ('Export Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ('Ratio Factor', str(comparison.ratio_factor)),
            ('', ''),
            ('Primary Survey', primary_file.file_name if primary_file else 'N/A'),
            ('Primary Survey Type', primary_file.survey_type if primary_file else 'N/A'),
            ('Reference Survey', reference_file.file_name if reference_file else 'N/A'),
            ('Reference Survey Type', reference_file.survey_type if reference_file else 'N/A'),
            ('', ''),
            ('--- STATISTICS ---', ''),
        ]

        # Add statistics
        stats = comparison.statistics
        metadata.extend([
            ('Point Count', str(stats.get('point_count', 0))),
            ('', ''),
            ('Max Deviation (Horizontal)', f"{stats.get('max_delta_horizontal', 0):.2f} m"),
            ('Max Deviation (Total 3D)', f"{stats.get('max_delta_total', 0):.2f} m"),
            ('MD at Max Horizontal Dev', f"{stats.get('md_at_max_horizontal', 0):.2f} m"),
            ('MD at Max Total Dev', f"{stats.get('md_at_max_total', 0):.2f} m"),
            ('', ''),
            ('Avg Horizontal Deviation', f"{stats.get('avg_delta_horizontal', 0):.2f} m"),
            ('Avg Total Deviation', f"{stats.get('avg_delta_total', 0):.2f} m"),
            ('Std Dev (Horizontal)', f"{stats.get('std_delta_horizontal', 0):.2f} m"),
            ('Std Dev (Total)', f"{stats.get('std_delta_total', 0):.2f} m"),
            ('', ''),
            ('Max ΔX', f"{stats.get('max_delta_x', 0):.2f} m"),
            ('Max ΔY', f"{stats.get('max_delta_y', 0):.2f} m"),
            ('Max ΔZ (TVD)', f"{stats.get('max_delta_z', 0):.2f} m"),
            ('', ''),
            ('Max ΔInclination', f"{stats.get('max_delta_inc', 0):.2f}°"),
            ('Max ΔAzimuth', f"{stats.get('max_delta_azi', 0):.2f}°"),
        ])

        # Write metadata rows
        for label, value in metadata:
            ws[f'A{row}'] = label
            if label:  # Only format non-empty labels
                ws[f'A{row}'].font = cls.METADATA_LABEL_FONT
            ws[f'B{row}'] = value
            ws[f'B{row}'].alignment = cls.METADATA_VALUE_ALIGNMENT
            row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25

    @classmethod
    def _write_position_deltas(cls, ws, comparison: ComparisonResult) -> None:
        """Write position deltas data to worksheet."""
        # Prepare data
        data = {
            'MD (m)': comparison.md_data,
            'ΔX (m)': comparison.delta_x,
            'ΔY (m)': comparison.delta_y,
            'ΔZ (m)': comparison.delta_z,
            'ΔHorizontal (m)': comparison.delta_horizontal,
            'ΔTotal 3D (m)': comparison.delta_total,
        }
        df = pd.DataFrame(data)

        # Write headers
        columns = list(data.keys())
        for col_num, column_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_name)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.alignment = cls.HEADER_ALIGNMENT

        # Write data rows with conditional formatting
        for row_num, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                if isinstance(value, (int, float)):
                    cell.number_format = '0.00'
                    # Highlight max values
                    if col_num > 1:  # Skip MD column
                        if abs(value) == max(abs(v) for v in data[columns[col_num-1]]):
                            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                cell.alignment = Alignment(horizontal='right')

        # Auto-adjust column widths
        for col_num in range(1, len(columns) + 1):
            ws.column_dimensions[chr(64 + col_num)].width = 16

    @classmethod
    def _write_angular_deltas(cls, ws, comparison: ComparisonResult) -> None:
        """Write angular deltas data to worksheet."""
        # Prepare data
        data = {
            'MD (m)': comparison.md_data,
            'ΔInclination (°)': comparison.delta_inc,
            'ΔAzimuth (°)': comparison.delta_azi,
        }
        df = pd.DataFrame(data)

        # Write headers
        columns = list(data.keys())
        for col_num, column_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_name)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.alignment = cls.HEADER_ALIGNMENT

        # Write data rows
        for row_num, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                if isinstance(value, (int, float)):
                    cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal='right')

        # Auto-adjust column widths
        for col_num in range(1, len(columns) + 1):
            ws.column_dimensions[chr(64 + col_num)].width = 16

    @classmethod
    def _write_survey_reference_data(cls, ws, comparison: ComparisonResult, survey_type: Literal['primary', 'reference']) -> None:
        """Write survey reference data (from calculated surveys)."""
        # Get the survey
        if survey_type == 'primary':
            survey_data = comparison.primary_survey
            title = 'Primary (Comparison) Survey Data'
        else:
            survey_data = comparison.reference_survey
            title = 'Reference Survey Data'

        # Title
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=12)

        # Get calculated data
        calc = survey_data.calculated_survey if hasattr(survey_data, 'calculated_survey') else None
        if not calc:
            ws['A3'] = 'Calculated data not available'
            return

        # Prepare data
        data = {
            'MD (m)': survey_data.md_data,
            'Inc (°)': survey_data.inc_data,
            'Azi (°)': survey_data.azi_data,
            'Easting (m)': calc.easting,
            'Northing (m)': calc.northing,
            'TVD (m)': calc.tvd,
        }
        df = pd.DataFrame(data)

        # Write headers
        columns = list(data.keys())
        for col_num, column_name in enumerate(columns, 1):
            cell = ws.cell(row=3, column=col_num, value=column_name)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.alignment = cls.HEADER_ALIGNMENT

        # Write data rows
        for row_num, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 4):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                if isinstance(value, (int, float)):
                    cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal='right')

        # Auto-adjust column widths
        for col_num in range(1, len(columns) + 1):
            ws.column_dimensions[chr(64 + col_num)].width = 14

    @classmethod
    def _export_comparison_csv(cls, comparison: ComparisonResult) -> Tuple[BinaryIO, str, str]:
        """Generate CSV file for comparison results (combined data)."""
        # Prepare combined data
        data = {
            'MD (m)': comparison.md_data,
            'ΔX (m)': comparison.delta_x,
            'ΔY (m)': comparison.delta_y,
            'ΔZ (m)': comparison.delta_z,
            'ΔHorizontal (m)': comparison.delta_horizontal,
            'ΔTotal 3D (m)': comparison.delta_total,
            'ΔInclination (°)': comparison.delta_inc,
            'ΔAzimuth (°)': comparison.delta_azi,
        }
        df = pd.DataFrame(data)

        # Create buffer with metadata header
        buffer = io.StringIO()

        # Write metadata as comments
        primary_file = comparison.primary_survey.survey_file
        reference_file = comparison.reference_survey.survey_file
        stats = comparison.statistics

        buffer.write(f"# Survey Comparison Export\n")
        buffer.write(f"# Run Name: {str(comparison.run)}\n")
        buffer.write(f"# Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        buffer.write(f"# Primary Survey: {primary_file.file_name if primary_file else 'N/A'}\n")
        buffer.write(f"# Reference Survey: {reference_file.file_name if reference_file else 'N/A'}\n")
        buffer.write(f"# Ratio Factor: {comparison.ratio_factor}\n")
        buffer.write(f"#\n")
        buffer.write(f"# Max Horizontal Deviation: {stats.get('max_delta_horizontal', 0):.2f} m\n")
        buffer.write(f"# Max Total Deviation: {stats.get('max_delta_total', 0):.2f} m\n")
        buffer.write(f"# Point Count: {stats.get('point_count', 0)}\n")
        buffer.write(f"#\n")

        # Write CSV data
        df.to_csv(buffer, index=False, float_format='%.2f')

        # Convert to bytes
        bytes_buffer = io.BytesIO(buffer.getvalue().encode('utf-8'))
        bytes_buffer.seek(0)

        # Generate filename
        filename = cls._generate_comparison_filename(
            comparison.run.run_name,
            primary_file.file_name if primary_file else 'primary',
            reference_file.file_name if reference_file else 'reference',
            format='csv'
        )

        return bytes_buffer, filename, 'text/csv'

    @classmethod
    def _generate_comparison_filename(
        cls,
        run_name: str,
        primary_name: str,
        reference_name: str,
        format: Literal['excel', 'csv']
    ) -> str:
        """
        Generate filename for comparison export.

        Format: {run_name}_comparison_{ref}_{comp}_{timestamp}.{ext}
        """
        # Sanitize names
        sanitized_run = re.sub(r'[^\w\s-]', '', run_name).strip()
        sanitized_run = re.sub(r'[-\s]+', '_', sanitized_run)[:30]

        sanitized_ref = re.sub(r'[^\w\s-]', '', reference_name).strip()
        sanitized_ref = re.sub(r'[-\s]+', '_', sanitized_ref)[:20]

        sanitized_comp = re.sub(r'[^\w\s-]', '', primary_name).strip()
        sanitized_comp = re.sub(r'[-\s]+', '_', sanitized_comp)[:20]

        # Generate timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        # Determine extension
        ext = 'xlsx' if format == 'excel' else 'csv'

        # Construct filename
        filename = f"{sanitized_run}_comparison_{sanitized_ref}_vs_{sanitized_comp}_{timestamp}.{ext}"

        return filename
