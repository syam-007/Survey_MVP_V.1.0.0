"""
Duplicate Survey API Views

Handles duplicate survey calculation (forward, inverse, comparison).
Results are NOT saved to database.
"""
import logging
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import io

from survey_api.services.duplicate_survey_service import DuplicateSurveyService
from survey_api.views.activity_log_viewset import log_activity
from survey_api.models import SurveyData

logger = logging.getLogger(__name__)


class DuplicateSurveyViewSet(viewsets.ViewSet):
    """ViewSet for duplicate survey operations."""

    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['post'], url_path='calculate')
    def calculate(self, request):
        """
        Calculate duplicate survey without saving to database.

        Request Body:
            {
                "survey_data_id": "uuid",
                "interpolation_step": 10.0
            }

        Response:
            200 OK: Duplicate survey data with forward, inverse, and comparison
            400 Bad Request: Validation error
            404 Not Found: Survey not found
            500 Internal Server Error: Calculation failed
        """
        try:
            survey_data_id = request.data.get('survey_data_id')
            interpolation_step = request.data.get('interpolation_step', 10.0)

            if not survey_data_id:
                return Response(
                    {'error': 'survey_data_id is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Validate interpolation step
            try:
                interpolation_step = float(interpolation_step)
                if interpolation_step <= 0:
                    raise ValueError("Interpolation step must be positive")
            except (ValueError, TypeError) as e:
                return Response(
                    {'error': f'Invalid interpolation_step: {str(e)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Calculate duplicate survey
            result = DuplicateSurveyService.calculate_duplicate_survey(
                survey_data_id=survey_data_id,
                interpolation_step=interpolation_step
            )

            # Log activity
            try:
                survey_data = SurveyData.objects.select_related('survey_file__run').get(id=survey_data_id)
                log_activity(
                    run_id=survey_data.survey_file.run.id,
                    user=request.user,
                    activity_type='duplicate_survey_calculated',
                    description=f'Calculated duplicate survey for {survey_data.survey_file.file_name} with interpolation step {interpolation_step}m',
                    metadata={
                        'survey_data_id': str(survey_data_id),
                        'survey_file_name': survey_data.survey_file.file_name,
                        'interpolation_step': interpolation_step,
                        'num_stations': len(result['forward_md'])
                    }
                )
            except Exception as log_error:
                logger.warning(f"Failed to log duplicate survey calculation: {str(log_error)}")

            return Response(result, status=status.HTTP_200_OK)

        except ValueError as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f"Duplicate survey calculation failed: {type(e).__name__}: {str(e)}")
            return Response(
                {'error': f'Failed to calculate duplicate survey: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'], url_path='export')
    def export(self, request):
        """
        Export duplicate survey results to Excel.

        Request Body:
            {
                "survey_data_id": "uuid",
                "interpolation_step": 10.0
            }

        Response:
            200 OK: Excel file download
            400 Bad Request: Validation error
            404 Not Found: Survey not found
            500 Internal Server Error: Export failed
        """
        try:
            survey_data_id = request.data.get('survey_data_id')
            interpolation_step = request.data.get('interpolation_step', 10.0)

            if not survey_data_id:
                return Response(
                    {'error': 'survey_data_id is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Calculate duplicate survey
            result = DuplicateSurveyService.calculate_duplicate_survey(
                survey_data_id=survey_data_id,
                interpolation_step=float(interpolation_step)
            )

            # Create Excel workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Header style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Sheet 1: Forward Results
            ws_forward = wb.create_sheet("Forward_Results")
            forward_headers = ['MD (m)', 'INC (deg)', 'AZI (deg)', 'North (m)', 'East (m)', 'TVD (m)']
            ws_forward.append(forward_headers)

            # Style headers
            for cell in ws_forward[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Add forward data
            for i in range(len(result['forward_md'])):
                ws_forward.append([
                    result['forward_md'][i],
                    result['forward_inc'][i],
                    result['forward_azi'][i],
                    result['forward_north'][i],
                    result['forward_east'][i],
                    result['forward_tvd'][i],
                ])

            # Auto-size columns
            for column in ws_forward.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws_forward.column_dimensions[column_letter].width = adjusted_width

            # Sheet 2: Inverse Results
            ws_inverse = wb.create_sheet("Inverse_Results")
            inverse_headers = ['MD (m)', 'INC (deg)', 'AZI (deg)', 'North (m)', 'East (m)', 'TVD (m)']
            ws_inverse.append(inverse_headers)

            # Style headers
            for cell in ws_inverse[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Add inverse data
            for i in range(len(result['forward_md'])):
                ws_inverse.append([
                    result['forward_md'][i],
                    result['inverse_inc'][i],
                    result['inverse_azi'][i],
                    result['inverse_north'][i],
                    result['inverse_east'][i],
                    result['inverse_tvd'][i],
                ])

            # Auto-size columns
            for column in ws_inverse.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws_inverse.column_dimensions[column_letter].width = adjusted_width

            # Sheet 3: Comparison
            ws_comparison = wb.create_sheet("Comparison")
            comparison_headers = [
                'MD (m)',
                'Delta INC (deg)', 'Delta AZI (deg)',
                'Delta North (m)', 'Delta East (m)', 'Delta TVD (m)',
                'Limit North', 'Limit East', 'Limit TVD'
            ]
            ws_comparison.append(comparison_headers)

            # Style headers
            for cell in ws_comparison[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Add comparison data
            for i in range(len(result['forward_md'])):
                ws_comparison.append([
                    result['forward_md'][i],
                    result['delta_inc'][i],
                    result['delta_azi'][i],
                    result['delta_north'][i],
                    result['delta_east'][i],
                    result['delta_tvd'][i],
                    result['limit_north'][i],
                    result['limit_east'][i],
                    result['limit_tvd'][i],
                ])

            # Auto-size columns
            for column in ws_comparison.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws_comparison.column_dimensions[column_letter].width = adjusted_width

            # Save to BytesIO
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            # Create response
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"duplicate_survey_{result['survey_file_name']}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            # Log activity
            try:
                survey_data = SurveyData.objects.select_related('survey_file__run').get(id=survey_data_id)
                log_activity(
                    run_id=survey_data.survey_file.run.id,
                    user=request.user,
                    activity_type='duplicate_survey_exported',
                    description=f'Exported duplicate survey results for {survey_data.survey_file.file_name}',
                    metadata={
                        'survey_data_id': str(survey_data_id),
                        'survey_file_name': survey_data.survey_file.file_name,
                        'interpolation_step': float(interpolation_step),
                        'filename': filename
                    }
                )
            except Exception as log_error:
                logger.warning(f"Failed to log duplicate survey export: {str(log_error)}")

            return response

        except ValueError as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f"Duplicate survey export failed: {type(e).__name__}: {str(e)}")
            return Response(
                {'error': f'Failed to export duplicate survey: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
