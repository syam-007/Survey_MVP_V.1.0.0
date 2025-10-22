"""
Unit tests for ExcelExportService.
"""
import io
from datetime import datetime
from unittest.mock import MagicMock, patch

import pandas as pd
from django.contrib.auth import get_user_model
from django.core.exceptions import ObjectDoesNotExist
from django.test import TestCase
from openpyxl import load_workbook

from survey_api.models import (
    CalculatedSurvey,
    Depth,
    InterpolatedSurvey,
    Location,
    Run,
    SurveyData,
    SurveyFile,
    TieOn,
)
from survey_api.services.excel_export_service import ExcelExportService


class ExcelExportServiceTest(TestCase):
    """Test cases for ExcelExportService."""

    def setUp(self):
        """Set up test data."""
        # Create user
        User = get_user_model()
        self.user = User.objects.create_user(
            username='testuser',
            email='test@example.com',
            password='testpass123'
        )

        # Create run
        self.run = Run.objects.create(
            run_number='RUN001',
            run_name='Test Run 001',
            run_type='GTL',
            user=self.user
        )

        # Create location
        self.location = Location.objects.create(
            run=self.run,
            latitude=29.5,
            longitude=-95.5
        )

        # Create depth
        self.depth = Depth.objects.create(
            run=self.run,
            reference_height=1000.0,
            reference_elevation=100.0
        )

        # Create tie-on
        self.tie_on = TieOn.objects.create(
            run=self.run,
            md=0.0,
            inc=0.0,
            azi=0.0,
            tvd=0.0,
            latitude=0.0,
            departure=0.0,
            well_type='Oil',
            hole_section='Production Casing',
            survey_tool_type='MWD',
            survey_interval_from=0.0,
            survey_interval_to=1000.0
        )

        # Create survey file
        self.survey_file = SurveyFile.objects.create(
            run=self.run,
            file_name='survey_data.xlsx',
            file_path='/uploads/survey_data.xlsx',
            file_size=102400,
            survey_type='GTL'
        )

        # Create survey data
        self.survey_data = SurveyData.objects.create(
            survey_file=self.survey_file,
            md_data=[0.0, 100.0, 200.0, 300.0],
            inc_data=[0.0, 5.0, 10.0, 15.0],
            azi_data=[0.0, 45.0, 90.0, 135.0],
            row_count=4
        )

        # Create calculated survey
        self.calculated = CalculatedSurvey.objects.create(
            survey_data=self.survey_data,
            easting=[0.0, 3.9, 15.5, 34.9],
            northing=[0.0, 3.9, 15.5, 34.9],
            tvd=[0.0, 99.9, 199.6, 298.9],
            dls=[0.0, 1.67, 1.67, 1.67],
            build_rate=[0.0, 1.67, 1.67, 1.67],
            turn_rate=[0.0, 0.0, 0.0, 0.0],
            calculation_status='calculated',
            calculation_duration=1.5,
            calculation_context={}
        )

        # Create interpolated survey
        self.interpolated = InterpolatedSurvey.objects.create(
            calculated_survey=self.calculated,
            resolution=10,
            md_interpolated=[0.0, 10.0, 20.0, 30.0, 40.0],
            inc_interpolated=[0.0, 0.5, 1.0, 1.5, 2.0],
            azi_interpolated=[0.0, 4.5, 9.0, 13.5, 18.0],
            easting_interpolated=[0.0, 0.39, 1.55, 3.49, 6.21],
            northing_interpolated=[0.0, 0.39, 1.55, 3.49, 6.21],
            tvd_interpolated=[0.0, 9.99, 19.96, 29.89, 39.78],
            dls_interpolated=[0.0, 0.5, 0.5, 0.5, 0.5],
            point_count=5,
            interpolation_status='completed',
            interpolation_duration=0.8
        )

    def test_export_calculated_survey_excel(self):
        """Test exporting calculated survey to Excel format."""
        buffer, filename, content_type = ExcelExportService.export_calculated_survey(
            str(self.calculated.id),
            format='excel'
        )

        # Verify return types
        self.assertIsInstance(buffer, io.BytesIO)
        self.assertIsInstance(filename, str)
        self.assertEqual(
            content_type,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # Verify filename format
        self.assertRegex(filename, r'^Test_Run_001_calculated_\d{8}_\d{6}\.xlsx$')

        # Load workbook and verify structure
        wb = load_workbook(buffer)
        self.assertIn('Metadata', wb.sheetnames)
        self.assertIn('Survey Data', wb.sheetnames)

        # Verify metadata sheet
        ws_meta = wb['Metadata']
        self.assertEqual(ws_meta['A1'].value, 'Survey Export Metadata')
        self.assertEqual(ws_meta['A3'].value, 'Run Name')
        self.assertEqual(ws_meta['B3'].value, 'RUN001 - Test Run 001')

        # Verify survey data sheet
        ws_data = wb['Survey Data']
        self.assertEqual(ws_data['A1'].value, 'MD (m)')
        self.assertEqual(ws_data['B1'].value, 'Inc (deg)')
        self.assertEqual(ws_data['A2'].value, 0.0)
        self.assertEqual(ws_data['B2'].value, 0.0)
        # 4 data rows + 1 header row = 5 rows
        self.assertEqual(ws_data.max_row, 5)

    def test_export_calculated_survey_csv(self):
        """Test exporting calculated survey to CSV format."""
        buffer, filename, content_type = ExcelExportService.export_calculated_survey(
            str(self.calculated.id),
            format='csv'
        )

        # Verify return types
        self.assertIsInstance(buffer, io.BytesIO)
        self.assertIsInstance(filename, str)
        self.assertEqual(content_type, 'text/csv')

        # Verify filename format
        self.assertRegex(filename, r'^Test_Run_001_calculated_\d{8}_\d{6}\.csv$')

        # Read CSV content
        buffer.seek(0)
        content = buffer.read().decode('utf-8')

        # Verify metadata comments
        self.assertIn('# Survey Export - Calculated Data', content)
        self.assertIn('# Run Name: RUN001 - Test Run 001', content)
        self.assertIn('# Row Count: 4', content)

        # Verify CSV headers
        self.assertIn('MD (m),Inc (deg),Azi (deg)', content)

        # Verify data rows
        lines = content.split('\n')
        data_lines = [line for line in lines if not line.startswith('#') and line.strip()]
        # 1 header + 4 data rows
        self.assertEqual(len(data_lines), 5)

    def test_export_interpolated_survey_excel(self):
        """Test exporting interpolated survey to Excel format."""
        buffer, filename, content_type = ExcelExportService.export_interpolated_survey(
            str(self.interpolated.id),
            format='excel'
        )

        # Verify return types
        self.assertIsInstance(buffer, io.BytesIO)
        self.assertIsInstance(filename, str)
        self.assertEqual(
            content_type,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # Verify filename format
        self.assertRegex(filename, r'^Test_Run_001_interpolated_\d{8}_\d{6}\.xlsx$')

        # Load workbook and verify structure
        wb = load_workbook(buffer)
        self.assertIn('Metadata', wb.sheetnames)
        self.assertIn('Survey Data', wb.sheetnames)

        # Verify metadata includes interpolation info
        ws_meta = wb['Metadata']
        # Find resolution row
        found_resolution = False
        for row in ws_meta.iter_rows(min_row=3, max_row=15):
            if row[0].value == 'Resolution (m)':
                self.assertEqual(row[1].value, '10')
                found_resolution = True
                break
        self.assertTrue(found_resolution, "Resolution not found in metadata")

        # Verify survey data sheet
        ws_data = wb['Survey Data']
        # 5 data rows + 1 header row = 6 rows
        self.assertEqual(ws_data.max_row, 6)

    def test_export_interpolated_survey_csv(self):
        """Test exporting interpolated survey to CSV format."""
        buffer, filename, content_type = ExcelExportService.export_interpolated_survey(
            str(self.interpolated.id),
            format='csv'
        )

        # Verify return types
        self.assertIsInstance(buffer, io.BytesIO)
        self.assertEqual(content_type, 'text/csv')

        # Read CSV content
        buffer.seek(0)
        content = buffer.read().decode('utf-8')

        # Verify metadata comments
        self.assertIn('# Survey Export - Interpolated Data', content)
        self.assertIn('# Resolution: 10 m', content)
        self.assertIn('# Point Count: 5', content)

    def test_export_nonexistent_calculated_survey(self):
        """Test exporting non-existent calculated survey raises error."""
        fake_id = '00000000-0000-0000-0000-000000000000'

        with self.assertRaises(ObjectDoesNotExist) as context:
            ExcelExportService.export_calculated_survey(fake_id)

        self.assertIn(fake_id, str(context.exception))

    def test_export_nonexistent_interpolated_survey(self):
        """Test exporting non-existent interpolated survey raises error."""
        fake_id = '00000000-0000-0000-0000-000000000000'

        with self.assertRaises(ObjectDoesNotExist) as context:
            ExcelExportService.export_interpolated_survey(fake_id)

        self.assertIn(fake_id, str(context.exception))

    def test_export_incomplete_calculated_survey(self):
        """Test exporting incomplete calculated survey raises error."""
        # Create another survey file for incomplete test
        survey_file2 = SurveyFile.objects.create(
            run=self.run,
            file_name='incomplete.xlsx',
            file_path='/uploads/incomplete.xlsx',
            file_size=1024,
            survey_type='GTL'
        )
        survey_data2 = SurveyData.objects.create(
            survey_file=survey_file2,
            md_data=[0.0],
            inc_data=[0.0],
            azi_data=[0.0],
            row_count=1
        )

        # Create incomplete calculated survey
        incomplete = CalculatedSurvey.objects.create(
            survey_data=survey_data2,
            easting=[],
            northing=[],
            tvd=[],
            dls=[],
            build_rate=[],
            turn_rate=[],
            calculation_status='pending',
            calculation_duration=0.0,
            calculation_context={}
        )

        with self.assertRaises(ValueError) as context:
            ExcelExportService.export_calculated_survey(str(incomplete.id))

        self.assertIn('pending', str(context.exception))

    def test_export_incomplete_interpolated_survey(self):
        """Test exporting incomplete interpolated survey raises error."""
        # Create incomplete interpolated survey with different resolution to avoid unique constraint
        incomplete = InterpolatedSurvey.objects.create(
            calculated_survey=self.calculated,
            resolution=5,  # Different resolution to avoid unique constraint
            md_interpolated=[],
            inc_interpolated=[],
            azi_interpolated=[],
            easting_interpolated=[],
            northing_interpolated=[],
            tvd_interpolated=[],
            dls_interpolated=[],
            point_count=0,
            interpolation_status='pending',
            interpolation_duration=0.0
        )

        with self.assertRaises(ValueError) as context:
            ExcelExportService.export_interpolated_survey(str(incomplete.id))

        self.assertIn('pending', str(context.exception))

    def test_export_calculated_missing_survey_data(self):
        """Test exporting calculated survey without survey_data - should not be possible due to FK constraint."""
        # This test is now obsolete because survey_data is a required OneToOne foreign key
        # CalculatedSurvey cannot exist without survey_data
        # Skipping this test case
        pass

    def test_generate_filename_sanitization(self):
        """Test filename sanitization removes special characters."""
        # Test with special characters
        filename = ExcelExportService._generate_filename(
            'Test@Run#001!',
            'calculated',
            'excel'
        )
        self.assertNotIn('@', filename)
        self.assertNotIn('#', filename)
        self.assertNotIn('!', filename)
        self.assertRegex(filename, r'^TestRun001_calculated_\d{8}_\d{6}\.xlsx$')

    def test_generate_filename_length_limit(self):
        """Test filename length is limited to 50 characters."""
        long_name = 'A' * 100
        filename = ExcelExportService._generate_filename(
            long_name,
            'calculated',
            'excel'
        )
        # Extract name part (before _calculated_)
        name_part = filename.split('_calculated_')[0]
        self.assertLessEqual(len(name_part), 50)

    def test_generate_filename_spaces_to_underscores(self):
        """Test filename converts spaces to underscores."""
        filename = ExcelExportService._generate_filename(
            'Test Run 001',
            'calculated',
            'csv'
        )
        self.assertNotIn(' ', filename)
        self.assertIn('Test_Run_001', filename)

    def test_calculated_columns_definition(self):
        """Test calculated columns are correctly defined."""
        expected_columns = [
            'MD (m)', 'Inc (deg)', 'Azi (deg)',
            'Easting (m)', 'Northing (m)', 'TVD (m)',
            'DLS (deg/30m)', 'Build Rate (deg/30m)', 'Turn Rate (deg/30m)'
        ]
        self.assertEqual(ExcelExportService.CALCULATED_COLUMNS, expected_columns)

    def test_interpolated_columns_definition(self):
        """Test interpolated columns are correctly defined."""
        expected_columns = [
            'MD (m)', 'Inc (deg)', 'Azi (deg)',
            'Easting (m)', 'Northing (m)', 'TVD (m)'
        ]
        self.assertEqual(ExcelExportService.INTERPOLATED_COLUMNS, expected_columns)

    def test_excel_header_styling(self):
        """Test Excel header cells have proper styling."""
        buffer, _, _ = ExcelExportService.export_calculated_survey(
            str(self.calculated.id),
            format='excel'
        )

        wb = load_workbook(buffer)
        ws = wb['Survey Data']

        # Check first header cell styling
        header_cell = ws['A1']
        self.assertTrue(header_cell.font.bold)
        # Color can be '00FFFFFF' or 'FFFFFFFF' depending on openpyxl version
        self.assertIn(header_cell.font.color.rgb, ['FFFFFFFF', '00FFFFFF'])
        # Fill color can be 'FF4472C4' or '004472C4' depending on openpyxl version
        self.assertIn(header_cell.fill.start_color.rgb, ['FF4472C4', '004472C4'])

    def test_excel_data_number_formatting(self):
        """Test Excel data cells have proper number formatting."""
        buffer, _, _ = ExcelExportService.export_calculated_survey(
            str(self.calculated.id),
            format='excel'
        )

        wb = load_workbook(buffer)
        ws = wb['Survey Data']

        # Check data cell formatting
        data_cell = ws['A2']
        self.assertEqual(data_cell.number_format, '0.00')

    def test_metadata_sheet_structure(self):
        """Test metadata sheet has proper structure and content."""
        buffer, _, _ = ExcelExportService.export_calculated_survey(
            str(self.calculated.id),
            format='excel'
        )

        wb = load_workbook(buffer)
        ws = wb['Metadata']

        # Verify title
        self.assertEqual(ws['A1'].value, 'Survey Export Metadata')
        self.assertTrue(ws['A1'].font.bold)

        # Verify run information
        labels = [cell.value for cell in ws['A'] if cell.row >= 3 and cell.value]
        self.assertIn('Run Name', labels)
        self.assertIn('Export Date', labels)
        self.assertIn('Data Type', labels)

    def test_csv_metadata_comments(self):
        """Test CSV includes metadata as comments."""
        buffer, _, _ = ExcelExportService.export_calculated_survey(
            str(self.calculated.id),
            format='csv'
        )

        buffer.seek(0)
        content = buffer.read().decode('utf-8')
        lines = content.split('\n')
        comment_lines = [line for line in lines if line.startswith('#')]

        # Verify multiple comment lines exist
        self.assertGreater(len(comment_lines), 3)

        # Verify specific metadata
        comment_text = '\n'.join(comment_lines)
        self.assertIn('Run Name:', comment_text)
        self.assertIn('Export Date:', comment_text)
